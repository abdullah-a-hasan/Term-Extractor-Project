import os
import csv
import json
from pathlib import Path


def read_csv_preview(csv_path: str, max_rows: int = 5) -> dict:
    """Read a CSV file and return preview info."""
    try:
        with open(csv_path, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            rows = []
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                rows.append(row)

        with open(csv_path, newline='', encoding='utf-8-sig') as f:
            row_count = sum(1 for _ in f) - 1

        return {
            "success": True,
            "headers": headers,
            "preview_rows": rows,
            "row_count": row_count,
            "columns": len(headers)
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def load_json_file(json_path: str) -> dict:
    """Load a JSON file and return its contents."""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {"success": True, "data": data}
    except Exception as e:
        return {"success": False, "error": str(e)}


def save_json_file(json_path: str, data: dict) -> dict:
    """Save data to a JSON file."""
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return {"success": True}
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_session_path(output_dir: str, unique_id: str) -> str:
    """Get the path for a session file."""
    return os.path.join(output_dir, f".{unique_id}.session.json")


def save_session(output_dir: str, unique_id: str, data: dict) -> dict:
    """Save session data to a .session.json file."""
    try:
        session_path = get_session_path(output_dir, unique_id)
        with open(session_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return {"success": True, "path": session_path}
    except Exception as e:
        return {"success": False, "error": str(e)}


def load_session(output_dir: str, unique_id: str) -> dict:
    """Load session data if it exists."""
    try:
        session_path = get_session_path(output_dir, unique_id)
        if not os.path.exists(session_path):
            return {"success": True, "data": None}
        with open(session_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {"success": True, "data": data}
    except Exception as e:
        return {"success": False, "error": str(e)}


def export_to_excel(terms: list, output_path: str) -> dict:
    """Export terms list to Excel using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Terms"

        headers = ["Source Term", "Target Term", "Status", "Count", "Hits", "Occ. Ratio"]
        ws.append(headers)

        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for term in terms:
            ws.append([
                term.get("source", ""),
                term.get("target", ""),
                term.get("status", ""),
                term.get("count", 0),
                term.get("hits", 0),
                term.get("occ_ratio", 0),
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(output_path)
        return {"success": True, "path": output_path}
    except Exception as e:
        return {"success": False, "error": str(e)}
